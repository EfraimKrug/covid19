import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
#########################################################

# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

dataCollection = dict()

countryColumn = 2
yearColumn = 3
matchColumn = 4
dataColumn = 5

matchValue = 'Emissions per capita (metric tons of carbon dioxide)'
newSpreadSheet = 'otherStatsAir.xlsx'

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openGatherBook():
    wb = load_workbook('./data/allData.xlsx')
    return wb

def openTrx(fname):
    wb = load_workbook(fname)
    return wb[wb.sheetnames[0]]

# def openTrx():
#     wb = load_workbook('./data/UNStats.xlsx')
#     return wb

def openOther():
    wb = load_workbook('./data/UNAirPolution001.xlsx')
    return wb

def writeTrx(sheet, country, pop, year, line):
    sheet.cell(row=line, column=1).value = country
    sheet.cell(row=line, column=2).value = year
    sheet.cell(row=line, column=3).value = pop

def getValue(sheet, line):
    return sheet.cell(row=line, column=dataColumn).value

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    i = 0
    # print([country, firstLine])
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=countryColumn).value
        matchVal = sheet.cell(row=ln, column=matchColumn).value
        if lineCountry == country.decode("UTF-8") and matchValue.decode("UTF-8") == matchVal:
            year = sheet.cell(row=ln, column=yearColumn).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=countryColumn).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=countryColumn).value is country:
            return [sheet.cell(row=ln, column=countryColumn).value, ln]
    return ['DONE', sheet.max_row]

######################################################################
# this list of files is built one at a time from the gather
# series of programs -
######################################################################
def getFileList():
    fileNameList = []
    path = '.'
    if len(sys.argv) == 2:
        path = sys.argv[1]

    files = os.listdir(path)
    for name in files:
        if ".xlsx" in name[-5:] and "other" in name[0:6]:
            fileNameList.append(name)

    return fileNameList

def processFile(fname):
    tag = fname.replace("other","").replace(".xlsx","")
    sheet = openTrx(fname)
    for lineItem in range(4, 245):
        stateName = sheet.cell(row=lineItem, column=1).value
        stateValue = sheet.cell(row=lineItem, column=3).value
        if stateName in dataCollection:
            dataCollection[stateName][tag] = stateValue
        else:
            dataCollection[stateName] = dict()
            dataCollection[stateName][tag] = stateValue

def gatherCollection(gatherSheet):
    gatherSheet.cell(row=1, column=3).value = "PopDensity"
    gatherSheet.cell(row=1, column=4).value = "LandUse"
    gatherSheet.cell(row=1, column=5).value = "Health001"
    gatherSheet.cell(row=1, column=6).value = "Health002"
    gatherSheet.cell(row=1, column=7).value = "Air"
    gatherSheet.cell(row=1, column=8).value = "Agriculture"
    for stateName in dataCollection:
        for r in range(2, 120):
            if gatherSheet.cell(row=r, column=1).value:
                if stateName == gatherSheet.cell(row=r, column=1).value:
                    if 'StatsPopDensity' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=3).value = str(dataCollection[stateName]['StatsPopDensity'])
                    if 'StatsLandUse' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=4).value = str(dataCollection[stateName]['StatsLandUse'])
                    if 'StatsHealth001' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=5).value = str(dataCollection[stateName]['StatsHealth001'])
                    if 'StatsHealth002' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=6).value = str(dataCollection[stateName]['StatsHealth002'])
                    if 'StatsAir' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=7).value = str(dataCollection[stateName]['StatsAir'])
                    if 'StatsAgriculture' in dataCollection[stateName]:
                        gatherSheet.cell(row=r, column=8).value = str(dataCollection[stateName]['StatsAgriculture'])

def printCollection():
    # sep = "\t"
    sep2 = "\t"
    for stateName in dataCollection:
        print(stateName)
        # sep = "\t"
        # if stateName and len(stateName) < 15:
        #     sep = "\t\t"
        for tag in dataCollection[stateName]:
            sep2 = "\t"
            if tag and len(tag) < 16:
                sep2 = "\t\t"
            if tag and len(tag) < 7:
                sep2 = "\t\t\t"

            print("\t\t\t" + str(tag) + str(sep2) + str(dataCollection[stateName][tag]))
#########################################################################
# Starting the program...
#########################################################################


for fileName in getFileList():
    # print(fileName)
    dataCollection[fileName] = dict()
    processFile(fileName)

gatherBook = openGatherBook()
gatherSheet = gatherBook[gatherBook.sheetnames[0]]
gatherCollection(gatherSheet)
# printCollection()
# print(dataCollection)

########################################################################
#
# trx = openTrx()
# other = openOther()
# countryLine = ['START', 3]
# popDensity = []
# trxLine = 4
#
# while countryLine[0] is not 'DONE':
#     countryLine = getNextCountry(other[other.sheetnames[0]], countryLine[1])
#     # if countryLine[0] is 'DONE':
#     #     break
#     # print(countryLine[0])
#     try:
#         popDensity = getCountryLines(other[other.sheetnames[0]], countryLine[0], countryLine[1])
#     except:
#         continue
#     # if not popDensity:
#     #     break
#     dataArray = getLatest(popDensity)
#     if dataArray[0] == 0:
#         continue
#     popData = getValue(other[other.sheetnames[0]], dataArray[0])
#     writeTrx(trx[trx.sheetnames[0]], countryLine[0], popData, dataArray[1], trxLine)
#     trxLine = trxLine + 1
#
# getTrx(trx[trx.sheetnames[0]], pops[pops.sheetnames[0]])
gatherBook.save('./data/allData.xlsx')
gatherBook.close()
# other.close()
