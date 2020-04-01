import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

countryColumn = 2
yearColumn = 3
matchColumn = 4
dataColumn = 5

matchValue = 'Health personnel: Physicians (per 1000 population)'
newSpreadSheet = 'otherStatsHealth002.xlsx'

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def outBook():
    wb = load_workbook('./data/allData.xlsx')
    return wb

def openPopulationBook():
    wb = load_workbook('./data/populations.xlsx')
    return wb

def openOther():
    wb = load_workbook('./data/UNHealth001.xlsx')
    return wb

def openTrx():
    wb = load_workbook('./data/time_series_covid_19_confirmed.xlsx')
    return wb

def writeTrx(sheet, country, pop, year, line):
    sheet.cell(row=line, column=1).value = country
    sheet.cell(row=line, column=2).value = year
    sheet.cell(row=line, column=3).value = pop

def getValue(sheet, line):
    return sheet.cell(row=line, column=dataColumn).value

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    i = 0
    # print([country, firstLine])
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=countryColumn).value
        matchVal = sheet.cell(row=ln, column=matchColumn).value
        if lineCountry == country.decode("UTF-8") and matchValue.decode("UTF-8") == matchVal:
            year = sheet.cell(row=ln, column=yearColumn).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=countryColumn).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=countryColumn).value is country:
            return [sheet.cell(row=ln, column=countryColumn).value, ln]
    return ['DONE', sheet.max_row]

########################################################################

trx = openTrx()
other = openOther()
outBook = outBook()
# countryLine = ['START', 3]
# popDensity = []
# trxLine = 4
sheet = trx[trx.sheetnames[0]]
oSheet = other[other.sheetnames[0]]
outSheet = outBook[outBook.sheetnames[0]]

# copy the date titles...
outSheet.cell(row=1, column=1).value = "Country"
outSheet.cell(row=1, column=2).value = "Population"

for t in range(5, 90):
    t1 = t + 5
    outSheet.cell(row=1, column=t1).value = sheet.cell(row=1, column=t).value

# get COVID data for relevant countries
for r in range(2, 500):
    if not sheet.cell(row=r, column=1).value and sheet.cell(row=r, column=2).value:
        for s in range(2, 2260):
            if oSheet.cell(row=s, column=2).value == sheet.cell(row=r, column=2).value:
                print(oSheet.cell(row=s, column=2).value)
                outSheet.cell(row=r, column=1).value = sheet.cell(row=r, column=2).value
                for t in range(5, 100):
                    t1 = t + 5
                    outSheet.cell(row=r, column=t1).value = sheet.cell(row=r, column=t).value
                # print("Last Column: " + str(t) + "/" + str(t1))
                # break

lastRow = 0
for s in range(3, 2500):
    r = 2503 - s
    if outSheet.cell(row=r, column=1).value:
        lastRow = r
        break

print("Last Row: " + str(lastRow))

popBook = openPopulationBook()
popSheet = popBook[popBook.sheetnames[0]]

for r in range(2, lastRow):
    country = outSheet.cell(row=r, column=1).value
    for s in range(2, lastRow):
        if country == popSheet.cell(row=s, column=1).value:
            outSheet.cell(row=r, column=2).value = popSheet.cell(row=s, column=2).value

print("Populations entered")

sw = False
for r in range(2, lastRow):
    if not outSheet.cell(row=r, column=1).value:
        if sw:
            last = r 
        else:
            sw = True
            first = r
            count = 1

# delete blank lines
# for s in range(3, lastRow):
#     r = lastRow - s + 3
#     if not outSheet.cell(row=r, column=1).value:
#         print("Deleting..." + str(r))
#         outSheet.delete_rows(r,1)
#
# print("Deleted empty rows...")
# lastRow = 0
# for s in range(3, 2500):
#     r = 2503 - s
#     if outSheet.cell(row=r, column=1).value:
#         lastRow = r
#         break
# print(lastRow)
# print("Last Row: " + str(lastRow))

# get population data

outBook.save("./data/allData.xlsx")
outBook.close()
trx.close()
other.close()
