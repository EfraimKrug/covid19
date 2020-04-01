import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openTrx():
    wb = load_workbook('./data/new.xlsx')
    return wb

def openPops():
    wb = load_workbook('./data/populations.xlsx')
    return wb

def getLast(sheet):
    dt = sheet.cell(row=1, column=sheet.max_column).value
    return dt

def getTrx(sheet, popsheet):
    for r in range(2, sheet.max_row):
        for s in range(1, popsheet.max_row):
            val01P   = sheet.cell(row=r, column=1).value
            val02P   = sheet.cell(row=r, column=2).value
            val01Q  = popsheet.cell(row=s, column=1).value
            if val02P == val01Q and (not val01P):
                sheet.cell(row=r, column=3).value = popsheet.cell(row=s, column=2).value

########################################################################

trx = openTrx()
pops = openPops()

getTrx(trx[trx.sheetnames[0]], pops[pops.sheetnames[0]])
trx.save("new.xlsx")
trx.close()
pops.close()
