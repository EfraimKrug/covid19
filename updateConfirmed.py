import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openDTrx():
    wb = load_workbook('./data/time_series_covid_19_deaths.xlsx')
    return wb

def openDUTrx():
    wb = load_workbook('../../Downloads/time_series_covid_19_deaths.xlsx')
    return wb

def openTrx():
    wb = load_workbook('./data/time_series_covid_19_confirmed.xlsx')
    return wb

def openUTrx():
    wb = load_workbook('../../Downloads/time_series_covid_19_confirmed.xlsx')
    return wb

def getLast(sheet):
    dt = sheet.cell(row=1, column=sheet.max_column).value
    return dt

def getTrx(sheet, usheet):
    startRange = sheet.max_column + 1
    endRange = usheet.max_column + 1
    # print("Start: " + str(startRange))
    # print("End: " + str(endRange))
    dU = getLast(usheet)
    d0 = getLast(sheet)

    global grid
    if str(d0) is str(dU):
        print ("all caught up")
        return
    # else:
    #     print("Sheet: " + str(d0))
    #     print("USheet: " + str(dU))

    p = q = 1
    # print("max_row: " + str(usheet.max_row))
    for r in range(1, sheet.max_row):
        for s in range(1, usheet.max_row):
            val01P   = sheet.cell(row=r, column=1).value
            val02P   = sheet.cell(row=r, column=2).value
            val01Q  = usheet.cell(row=s, column=1).value
            val02Q  = usheet.cell(row=s, column=2).value
            if val01P == val01Q and val02P == val02Q:
                for i in range(startRange, endRange):
                    sheet.cell(row=r, column=i).value = usheet.cell(row=s, column=i).value
                    # print("copying: [" + str(q) + "," + str(j) + "][" + str(p) + "," + str(j) + "]")

    # for r in range(1, usheet.max_row):
    #     val01P   = sheet.cell(row=p, column=1).value
    #     val02P   = sheet.cell(row=p, column=2).value
    #     val01Q  = usheet.cell(row=q, column=1).value
    #     val02Q  = usheet.cell(row=q, column=2).value
    #
    #     if val01P == val01Q and val02P == val02Q:
    #         print("starting the copy - maybe")
    #         print(val01P + "/" + val02P)
    #         for j in range(startRange, endRange):
    #             sheet.cell(row=p, column=j).value = usheet.cell(row=q, column=j).value
    #             print("copying: [" + str(q) + "," + str(j) + "][" + str(p) + "," + str(j) + "]")
    #         # print("Match: " + str(val01P) + " " + str(val02P))
    #     else:
    #         for i in range(q+1, q+5):
    #             val01Pt   = sheet.cell(row=p, column=1).value
    #             val02Pt   = sheet.cell(row=p, column=2).value
    #             val01Qt  = usheet.cell(row=i, column=1).value
    #             val02Qt  = usheet.cell(row=i, column=2).value
    #             # print(str(val01Pt) + "//" + str(val01Qt))
    #             if val01Pt == val01Qt and val02Pt == val02Qt:
    #                 p = i
    #                 break
    #     # print("Row being processed: " + str(r))
    #     p = p + 1
    #     q = q + 1
    #

trx = openTrx()
utrx = openUTrx()

getTrx(trx[trx.sheetnames[0]], utrx[utrx.sheetnames[0]])
trx.save("new.xlsx")
trx.close()
utrx.close()

trx = openDTrx()
utrx = openDUTrx()
getTrx(trx[trx.sheetnames[0]], utrx[utrx.sheetnames[0]])
trx.save("newD.xlsx")
trx.close()
utrx.close()
