import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openTrx():
    wb = load_workbook('./data/UNStats.xlsx')
    return wb

def openOther():
    wb = load_workbook('./data/UNPopulation001.xlsx')
    return wb

def writeTrx(sheet, country, pop, year, line):
    sheet.cell(row=line, column=1).value = country
    sheet.cell(row=line, column=2).value = year
    sheet.cell(row=line, column=3).value = pop

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    val = 'Population density'
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=2).value
        matchVal = sheet.cell(row=ln, column=4).value
        if lineCountry == country and val == matchVal:
            year = sheet.cell(row=ln, column=3).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=2).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=2).value is country:
            return [sheet.cell(row=ln, column=2).value, ln]
    return ['DONE', sheet.max_row]

def getTrx(sheet, popsheet):
    for r in range(2, sheet.max_row):
        for s in range(1, popsheet.max_row):
            val01P   = sheet.cell(row=r, column=1).value
            val02P   = sheet.cell(row=r, column=2).value
            val01Q  = popsheet.cell(row=s, column=1).value
            if val02P == val01Q and (not val01P):
                sheet.cell(row=r, column=3).value = popsheet.cell(row=s, column=2).value

########################################################################

trx = openTrx()
other = openOther()
countryLine = ['START', 3]
popDensity = []
trxLine = 1

while countryLine[0] is not 'DONE':
    countryLine = getNextCountry(other[other.sheetnames[0]], countryLine[1])
    popDensity = getCountryLines(other[other.sheetnames[0]], countryLine[0], countryLine[1])
    dataArray = getLatest(popDensity)
    writeTrx(trx[trx.sheetnames[0]], countryLine[0], dataArray[0], dataArray[1], trxLine)
    trxLine = trxLine + 1
    # print(countryLine[0])

# getTrx(trx[trx.sheetnames[0]], pops[pops.sheetnames[0]])
trx.save("otherStats.xlsx")
trx.close()
other.close()
