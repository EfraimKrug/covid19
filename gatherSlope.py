import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
#########################################################

# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

dataCollection = dict()

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openGatherBook():
    wb = load_workbook('./data/allData.xlsx')
    return wb

def getValue(sheet, line):
    return sheet.cell(row=line, column=dataColumn).value

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    i = 0
    # print([country, firstLine])
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=countryColumn).value
        matchVal = sheet.cell(row=ln, column=matchColumn).value
        if lineCountry == country.decode("UTF-8") and matchValue.decode("UTF-8") == matchVal:
            year = sheet.cell(row=ln, column=yearColumn).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=countryColumn).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=countryColumn).value is country:
            return [sheet.cell(row=ln, column=countryColumn).value, ln]
    return ['DONE', sheet.max_row]

######################################################################
# this list of files is built one at a time from the gather
# series of programs -
######################################################################
def getFileList():
    fileNameList = []
    path = '.'
    if len(sys.argv) == 2:
        path = sys.argv[1]

    files = os.listdir(path)
    for name in files:
        if ".xlsx" in name[-5:] and "other" in name[0:6]:
            fileNameList.append(name)

    return fileNameList

def processFile(fname):
    tag = fname.replace("other","").replace(".xlsx","")
    sheet = openTrx(fname)
    for lineItem in range(4, 245):
        stateName = sheet.cell(row=lineItem, column=1).value
        stateValue = sheet.cell(row=lineItem, column=3).value
        if stateName in dataCollection:
            dataCollection[stateName][tag] = stateValue
        else:
            dataCollection[stateName] = dict()
            dataCollection[stateName][tag] = stateValue

def getSlope(sheet, row):
    values = []
    retValues = []
    for c in range(14, 100):

        if sheet.cell(row=row, column=c).value and int(sheet.cell(row=row, column=c).value) > 0:
            values.append(sheet.cell(row=row, column=c).value)
    retValues = [len(values), values[-1]/len(values)]
    return retValues

def gatherSlopes(gatherSheet):
    gatherSheet.cell(row=1, column=13).value = "Slope"
    for r in range(2, 270):
        if gatherSheet.cell(row=r, column=1).value:
            # stateName = gatherSheet.cell(row=r, column=1).value
            # dataCollection[stateName] = getSlope(gatherSheet, r)
            gatherSheet.cell(row=r, column=13).value = getSlope(gatherSheet, r)[1]
                # gatherSheet.cell(row=r, column=12).value = str(dataCollection[stateName]['StatsWater'])

def printCollection():
    for stateName in dataCollection:
        print(stateName + ":" + str(dataCollection[stateName][0]) + ":" + str(dataCollection[stateName][1]))

#########################################################################
# Starting the program...
#########################################################################

gatherBook = openGatherBook()
gatherSheet = gatherBook[gatherBook.sheetnames[0]]
gatherSlopes(gatherSheet)
printCollection()
gatherBook.save('./data/allData.xlsx')
gatherBook.close()
# other.close()
