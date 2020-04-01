import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
##########################################################################
#   Stats from Influenza in America:
#   Population: 327,200,000
#       Highest estimates/Lowest estimates
#       45,000,000/9,300,000 affected (13.75%/2.84% of the population)
#          810,000/140,000 hospitalized
#           61,000/12,000 deaths (.0186%/.0037% of the population)
##########################################################################
# get parent directory...
sys.path.append(os.getcwd())
sys.path.append(os.getcwd()[0:os.getcwd().rfind('\\')])

import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

SHOW_SWITCH = True

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

cummData = []
diffData = []
percData = []
perc2Data = []
dateData = []
dData = []
maxLine = []

def openTrx():
    #wb = load_workbook('./data/time_series_covid_19_confirmed.xlsx')
    wb = load_workbook('./data/new.xlsx')
    return wb

def openDTrx():
    wb = load_workbook('./data/newD.xlsx')
    # wb = load_workbook('./data/time_series_covid_19_deaths.xlsx')
    return wb

def getMaxUS(sheet):
    valList = []
    rowCollection = []

    lastCol = 0
    for r in range(2, sheet.max_column):
        val = sheet.cell(row=5, column=r).value
        if val:
            lastCol = r

    for r in range(2, sheet.max_row):
        country = sheet.cell(row=r, column=2).value
        if country == "US" and not sheet.cell(row=r, column=1).value:
            val = sheet.cell(row=r, column=lastCol).value
            valList.append([r,val])

    for i in range(0, len(valList)):
        checkVal = 0
        checkRow = 0
        for j in range(0,len(valList)):
            if (valList[j][1] > checkVal) and (valList[j][0] not in rowCollection):
                checkVal = valList[j][1]
                checkRow = valList[j][0]

        rowCollection.append(checkRow)

    return rowCollection

def getMaxConfirmed(sheet):
    valList = []
    rowCollection = []

    lastCol = 0
    for r in range(2, sheet.max_column):
        if sheet.cell(row=5, column=r).value:
            lastCol = r

    for r in range(2, sheet.max_row):
        val = sheet.cell(row=r, column=lastCol).value
        valList.append([r,val])

    for i in range(0,15):
        checkVal = 0
        checkRow = 0
        for j in range(0,len(valList)):
            if (valList[j][1] > checkVal) and (valList[j][0] not in rowCollection):
                checkVal = valList[j][1]
                checkRow = valList[j][0]
        rowCollection.append(checkRow)
    return rowCollection


def getTrx(sheet, dsheet, r_in):
    global title
    global cummData
    global diffData
    global percData
    global perc2Data
    global dData
    global dateData
    global maxLine

    cummData = []
    diffData = []
    percData = []
    perc2Data = []
    dData = []
    dateData = []
    maxLine = []
    MassPop = 82000000

    #prevVal = 1
    cummValue = prevVal = difference = perc = perc2 = dVal = 0
    rVal = r_in
    # population = 82000000
    try:
        population = int(sheet.cell(row=rVal, column=3).value)
    except:
        population = 99999999
    maxL = float(.0002) * float(population)
    for r in range(5, sheet.max_column):
        # print([r,rVal])
        if sheet.cell(row=rVal, column=r).value > 0:
            fld1 = str(sheet.cell(row=1, column=r).value)[:-3]
            cummValue = sheet.cell(row=rVal, column=r).value
            fld2 = "\t\t" + str(cummValue)
            difference = int(sheet.cell(row=rVal, column=r).value) - prevVal
            fld3 = "\t\t" + str(difference)
            perc = 0
            if prevVal > 0:
                perc = int(difference * 100) / int(prevVal)
            perc2 = float(cummValue * 100) / float(population)
            sperc2 = str(round(perc2,4))
            fld4 = "\t\t" + str(perc) + "%"
            fld5 = "\t\t" + sperc2 + "%"
            # print(fld1 + fld2 + fld5 + fld3 + fld4)
            prevVal = sheet.cell(row=rVal, column=r).value
            if sheet.cell(row=rVal, column=1).value == dsheet.cell(row=rVal, column=1).value:
                if sheet.cell(row=rVal, column=2).value == dsheet.cell(row=rVal, column=2).value:
                    dVal = dsheet.cell(row=rVal, column=r).value

            dateData.append(fld1)
            # dateData.append(r)
            cummData.append(cummValue)
            diffData.append(difference)
            percData.append(perc)
            perc2Data.append(sperc2)
            dData.append(dVal)
            maxLine.append(maxL)

    popPerc = float(cummValue * 100) / float(population)
    spopPerc = str(round(popPerc,2))
    country = sheet.cell(row=rVal, column=2).value
    dPop = float(dVal * 100) / float(cummValue)
    dBPop = float(dVal * 100) / float(population)
    if(len(str(sheet.cell(row=rVal, column=1).value)) > 4):
        country = sheet.cell(row=rVal, column=2).value + ", " + str(sheet.cell(row=rVal, column=1).value)
    title = country + " " + spopPerc + "% pop & " + str(round(dPop,2)) + "% affected d..."
    likly01 = float(MassPop) * float(spopPerc) / float(MassPop)
    likly02 = float(likly01) * float(dBPop)
    pDPop = float(dVal) * float(100) / float(population)
    suptitle = str(population) + ", " + str(cummValue) + " / " + str(dVal) + " (p,c/d) [" + str(round(pDPop,4)) + "%]"
    # suptitle = str(MassPop) + ":" + str(spopPerc)
    # suptitle = "Mass: " + str(round(likly01,5)) + " chances & " + str(round(likly02,5))
    #+ " or 1:"  + str(round(likly02,3))
    writeHTML("<td>" + str(country) + "</td><td>" + str(population) + "</td><td>" + str(cummValue) + "</td><td>[" + str(round(dPop,4)) + "%]</td><td>" + str(dVal) + "</td><td>[" + str(round(pDPop,4)) + "%]</td>")

    if SHOW_SWITCH:
        df = ""
        df = pd.DataFrame({
                'dates':dateData,
                'cumm':cummData,
                'diff':diffData,
                'perc':percData,
                'perc2':perc2Data,
                'dData':dData,
                'maxLine':maxLine
        })

        ax = plt.gca()

        df.plot(kind='line', x='dates', y='cumm', ax=ax)
        df.plot(kind='line', x='dates', y='dData', color='blue', ax=ax)
        df.plot(kind='line', x='dates', y='maxLine', color='yellow', ax=ax)

        df.set_index('dates', drop=True, append=True, inplace=True, verify_integrity=False)
        plt.title(title)
        plt.suptitle(suptitle)
        plt.show()


def rCountries():
    for i in [2,3,4,13,15,16,17,18,35]:
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany

def rChina():
    for i in range(164,189):
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany

# state by state
def rUS():
    for i in range(102,106):
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany
    for i in range(108,137):
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany
    for i in range(141,158):
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany

def rInterest():
    for i in [13,18,35,102,103,104,105,159]:
        getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)  # germany

def getInteresting():
    done = []
    # rows = getMaxUS(trx[trx.sheetnames[0]])
    # print(rows)
    # for i in rows:
    #     if i not in done:
    #         print(i)
    #         getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)
    #         done.append(i)

    rows = getMaxConfirmed(trx[trx.sheetnames[0]])
    for i in rows:
        if i not in done:
            getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)
            done.append(i)
    # highest death rates
    rows = getMaxConfirmed(dtrx[dtrx.sheetnames[0]])
    for i in rows:
        if i not in done:
            getTrx(trx[trx.sheetnames[0]], dtrx[dtrx.sheetnames[0]], i)
            done.append(i)

def firstWrite():
    str = "<p><br>   Stats from Influenza in America:"
    str = str + "<br>   Population: 327,200,000"
    str = str + "<br>       Highest estimates/Lowest estimates"
    str = str + "<br>       45,000,000/9,300,000 affected (13.75%/2.84% of the population)"
    str = str + "<br>          810,000/140,000 hospitalized"
    str = str + "<br>           61,000/12,000 deaths (.0186%/.0037% of the population)"
    str = str + "<br>"
    str = str + "</p></div><div id=bottom>"

    with open("./data/stats.html","w+") as f:
        f.write("<html><head><title>Carona Stats</title></head><body>")
        f.write("<style>body {margin:3%;background:black;width:70%;} table {margin: 5%;} table, th, td {border: 2px solid black;padding:4px;} th{color:blue;background:lightblue;} #top{color:beige;background:grey;margin:5%;padding:5%;font-size:1.5em;}</style>")
        f.write("<div id=top><table>")
        f.write("<tr><th>Country</th><th>Population</th><th>Confirmed</th><th>% of population</th><th>Deaths</th><th>% of population</th></tr>")
        f.write(str)
        f.close()

def writeHTML(text):
    with open("./data/stats.html","a+") as f:
        f.write("<tr>" + text + "</tr>")
        f.close()

def lastWrite():
    graphs = "<table><tr><td><img src='./images/italy.jpeg'/></td><td><img src='./images/china.jpeg'/></td></tr><tr><td><img src='./images/spain.jpeg'/></td><td><img src='./images/germany.jpeg'/></td></tr><tr><td><img src='./images/iran.jpeg'/></td><td><img src='./images/switzerland.jpeg'/></td></tr><tr><td><img src='./images/korea.jpeg'/></td><td><img src='./images/netherlands.jpeg'/></td></tr><tr><td><img src='./images/austria.jpeg'/></td><td><img src='./images/belgium.jpeg'/></td></tr><tr><td><img src='./images/turkey.jpeg'/></td><td><img src='./images/portugal.jpeg'/></td></tr><tr><td><img src='./images/norway.jpeg'/></td><td><img src='./images/brazil.jpeg'/></td></tr><tr><td><img src='./images/sweden.jpeg'/></td><td><img src='./images/indonesia.jpeg'/></td></tr></table>"

    with open("./data/stats.html","a+") as f:
        f.write("</table></div>")
        f.write(graphs)
        f.write("</body></html>")


firstWrite()
trx = openTrx()
dtrx = openDTrx()
# getInteresting()
rInterest()
rCountries()
rChina()
# rUS()
lastWrite()
trx.close()
