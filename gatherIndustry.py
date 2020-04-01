import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

countryColumn = 2
yearColumn = 4
matchColumn = 3
dataColumn = 5

matchValue = 'Index of industrial production: Total industry - Mining; manufacturing; electricity, gas and water (Index base: 2005=100)'
newSpreadSheet = 'otherStatsIndustry.xlsx'

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openTrx():
    wb = load_workbook('./data/UNStats.xlsx')
    return wb

def openOther():
    wb = load_workbook('./data/UNIndustry001.xlsx')
    return wb

def writeTrx(sheet, country, pop, year, line):
    sheet.cell(row=line, column=1).value = country
    sheet.cell(row=line, column=2).value = year
    sheet.cell(row=line, column=3).value = pop

def getValue(sheet, line):
    return sheet.cell(row=line, column=dataColumn).value

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    i = 0
    # print([country, firstLine])
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=countryColumn).value
        matchVal = sheet.cell(row=ln, column=matchColumn).value
        if lineCountry == country.decode("UTF-8") and matchValue.decode("UTF-8") == matchVal:
            year = sheet.cell(row=ln, column=yearColumn).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=countryColumn).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=countryColumn).value is country:
            return [sheet.cell(row=ln, column=countryColumn).value, ln]
    return ['DONE', sheet.max_row]

########################################################################

trx = openTrx()
other = openOther()
countryLine = ['START', 3]
popDensity = []
trxLine = 4

while countryLine[0] is not 'DONE':
    countryLine = getNextCountry(other[other.sheetnames[0]], countryLine[1])
    # if countryLine[0] is 'DONE':
    #     break
    # print(countryLine[0])
    try:
        popDensity = getCountryLines(other[other.sheetnames[0]], countryLine[0], countryLine[1])
    except:
        continue
    # if not popDensity:
    #     break
    dataArray = getLatest(popDensity)
    if dataArray[0] == 0:
        continue
    popData = getValue(other[other.sheetnames[0]], dataArray[0])
    writeTrx(trx[trx.sheetnames[0]], countryLine[0], popData, dataArray[1], trxLine)
    trxLine = trxLine + 1

# getTrx(trx[trx.sheetnames[0]], pops[pops.sheetnames[0]])
trx.save(newSpreadSheet)
trx.close()
other.close()
