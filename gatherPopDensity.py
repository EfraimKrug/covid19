import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

#########################################################
# get parent directory...
import sys
import csv
from datetime import datetime
from datetime import time
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import get_column_letter

from openpyxl.styles.borders import Border, Side

import smtplib
from Profile import *
from periodProcess import *

toaddrs  = 'EfraimMKrug@gmail.com'
title = ''
# accountArray = []
grid = []
# accountArray.append(dict())
# server = 0
now_dt = datetime.today()
td = timedelta(days=-10)

def openTrx():
    wb = load_workbook('./data/UNStats.xlsx')
    return wb

def openOther():
    wb = load_workbook('./data/UNPopulation001.xlsx')
    return wb

def writeTrx(sheet, country, pop, year, line):
    sheet.cell(row=line, column=1).value = country
    sheet.cell(row=line, column=2).value = year
    sheet.cell(row=line, column=3).value = pop

def getValue(sheet, line):
    return sheet.cell(row=line, column=5).value

def getLatest(yearNumberList):
    year = 0
    number = 0
    for yn in yearNumberList:
        # print(yn)
        if yn[1] > year:
            year = yn[1]
            number = yn[0]

    return [number, year]

def getCountryLines(sheet, country, firstLine):
    lineList = []
    val = 'Population density'
    for ln in range(firstLine, sheet.max_row):
        lineCountry = sheet.cell(row=ln, column=2).value
        matchVal = sheet.cell(row=ln, column=4).value
        if lineCountry == country and val == matchVal:
            year = sheet.cell(row=ln, column=3).value
            lineList.append([ln, year])

    return lineList

def getNextCountry(sheet, firstLine):
    country = sheet.cell(row=firstLine, column=2).value
    for ln in range(firstLine, sheet.max_row):
        if not sheet.cell(row=ln, column=2).value is country:
            return [sheet.cell(row=ln, column=2).value, ln]
    return ['DONE', sheet.max_row]

########################################################################

trx = openTrx()
other = openOther()
countryLine = ['START', 3]
popDensity = []
trxLine = 4

while countryLine[0] is not 'DONE':
    countryLine = getNextCountry(other[other.sheetnames[0]], countryLine[1])
    if countryLine[0] is 'DONE':
        break
    popDensity = getCountryLines(other[other.sheetnames[0]], countryLine[0], countryLine[1])
    dataArray = getLatest(popDensity)
    popData = getValue(other[other.sheetnames[0]], dataArray[0])
    writeTrx(trx[trx.sheetnames[0]], countryLine[0], popData, dataArray[1], trxLine)
    trxLine = trxLine + 1
    # print(countryLine[0])

# getTrx(trx[trx.sheetnames[0]], pops[pops.sheetnames[0]])
trx.save("otherStats.xlsx")
trx.close()
other.close()
